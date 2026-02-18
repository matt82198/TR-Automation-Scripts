"""
QuickBooks Invoice Generator - Excel Export

Generates Excel workbooks with invoice line items ready for copy-paste
into QuickBooks Desktop Enterprise. Combines order data, payment matching,
and product mapping into a single downloadable file.
"""

from io import BytesIO
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers


def generate_invoice_excel(
    orders_raw: List[Dict],
    payment_results: List[Dict],
    sku_mapper,
    customer_matcher,
    ship_from_state: str = 'GA'
) -> BytesIO:
    """
    Generate an Excel workbook with invoice data for QuickBooks Desktop.

    Args:
        orders_raw: Raw order dicts from Squarespace API
        payment_results: Payment matching results from PaymentMatcher.match_orders()
        sku_mapper: ProductMapper instance (loaded with mappings)
        customer_matcher: CustomerMatcher instance (loaded with customers), or None
        ship_from_state: Business state for tax determination

    Returns:
        BytesIO containing the Excel workbook
    """
    from squarespace_to_quickbooks import (
        parse_variant_options, is_in_state_order,
        format_date_for_qb, sanitize_customer_name
    )

    wb = Workbook()

    # Build payment lookup by order number
    payment_lookup = {}
    for pr in payment_results:
        payment_lookup[str(pr['order_number'])] = pr

    # Build order lookup by order number
    order_lookup = {}
    for order in orders_raw:
        order_num = str(order.get('orderNumber', ''))
        order_lookup[order_num] = order

    # Resolve customer names and track new customers
    customer_names = {}  # order_number -> customer_name
    new_customers = {}   # customer_name -> customer_info

    for order in orders_raw:
        order_num = str(order.get('orderNumber', ''))
        billing = order.get('billingAddress', {})
        first_name = billing.get('firstName', '').strip()
        last_name = billing.get('lastName', '').strip()
        email = order.get('customerEmail', '').strip()
        phone = billing.get('phone', '').strip()

        matched_name = None
        if customer_matcher:
            matched_name = customer_matcher.find_match(email, phone, first_name, last_name)

        if matched_name:
            customer_names[order_num] = matched_name
        else:
            if first_name and last_name:
                name = f"{first_name} {last_name}"
            elif first_name or last_name:
                name = first_name or last_name
            else:
                name = email.split('@')[0] if email else 'Guest Customer'
            name = sanitize_customer_name(name)
            customer_names[order_num] = name

            if name not in new_customers:
                shipping = order.get('shippingAddress') or {}
                ship_first = (shipping.get('firstName', '') or '').strip()
                ship_last = (shipping.get('lastName', '') or '').strip()
                ship_name = f"{ship_first} {ship_last}".strip() or name

                bill_line1 = (billing.get('address1', '') or '').strip()
                bill_line2 = (billing.get('address2', '') or '').strip()
                bill_city = (billing.get('city', '') or '').strip()
                bill_state = (billing.get('state', '') or '').strip()
                bill_zip = (billing.get('postalCode', '') or '').strip()
                bill_addr = f"{bill_line1}, {bill_city}, {bill_state} {bill_zip}".strip(', ')
                if bill_line2:
                    bill_addr = f"{bill_line1}, {bill_line2}, {bill_city}, {bill_state} {bill_zip}".strip(', ')

                ship_line1 = (shipping.get('address1', '') or '').strip()
                ship_line2 = (shipping.get('address2', '') or '').strip()
                ship_city = (shipping.get('city', '') or '').strip()
                ship_state_val = (shipping.get('state', '') or '').strip()
                ship_zip = (shipping.get('postalCode', '') or '').strip()
                ship_addr = f"{ship_name}, {ship_line1}, {ship_city}, {ship_state_val} {ship_zip}".strip(', ')
                if ship_line2:
                    ship_addr = f"{ship_name}, {ship_line1}, {ship_line2}, {ship_city}, {ship_state_val} {ship_zip}".strip(', ')

                is_in_state = is_in_state_order(order, ship_from_state)
                new_customers[name] = {
                    'email': email,
                    'phone': phone,
                    'bill_address': bill_addr,
                    'ship_address': ship_addr,
                    'tax_code': 'Tax' if is_in_state else 'Non'
                }

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    separator_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    currency_fmt = '$#,##0.00'

    # =========================================================================
    # Sheet 1: Summary
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = [
        'Order #', 'Date', 'Customer', 'Gross', 'Payment Source',
        'Net', 'Processing Fee', 'Write-off', 'Status'
    ]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Use payment_results order (preserves user's input order)
    row = 2
    for pr in payment_results:
        order_num = str(pr['order_number'])
        customer = customer_names.get(order_num, pr.get('customer_name', ''))

        ws_summary.cell(row=row, column=1, value=order_num)
        ws_summary.cell(row=row, column=2, value=pr.get('order_date', ''))
        ws_summary.cell(row=row, column=3, value=customer)

        gross_cell = ws_summary.cell(row=row, column=4, value=pr.get('gross_amount', 0))
        gross_cell.number_format = currency_fmt

        ws_summary.cell(row=row, column=5, value=pr.get('payment_source', '') or '')

        net_cell = ws_summary.cell(row=row, column=6, value=pr.get('net_amount') or 0)
        net_cell.number_format = currency_fmt

        fee_cell = ws_summary.cell(row=row, column=7, value=pr.get('processing_fee') or 0)
        fee_cell.number_format = currency_fmt

        wo_cell = ws_summary.cell(row=row, column=8, value=pr.get('write_off') or 0)
        wo_cell.number_format = currency_fmt

        ws_summary.cell(row=row, column=9, value='Matched' if pr.get('matched') else 'Unmatched')
        row += 1

    # Totals row
    if payment_results:
        totals_row = row
        ws_summary.cell(row=totals_row, column=3, value='TOTALS').font = Font(bold=True)

        for col_idx, key in [(4, 'gross_amount'), (6, 'net_amount'), (7, 'processing_fee'), (8, 'write_off')]:
            total = sum(pr.get(key) or 0 for pr in payment_results)
            cell = ws_summary.cell(row=totals_row, column=col_idx, value=total)
            cell.number_format = currency_fmt
            cell.font = Font(bold=True)

    ws_summary.freeze_panes = 'A2'
    _auto_fit_columns(ws_summary)

    # =========================================================================
    # Sheet 2: Line Items
    # =========================================================================
    ws_lines = wb.create_sheet("Line Items")

    line_headers = [
        'Order #', 'Customer', 'Invoice Date', 'Item', 'Description',
        'Qty', 'Rate', 'Amount', 'Tax Code'
    ]
    for col, header in enumerate(line_headers, 1):
        cell = ws_lines.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for pr in payment_results:
        order_num = str(pr['order_number'])
        order = order_lookup.get(order_num)
        if not order:
            continue
        if order.get('fulfillmentStatus') == 'CANCELED':
            continue

        customer = customer_names.get(order_num, pr.get('customer_name', ''))
        invoice_date = format_date_for_qb(order.get('createdOn', ''))
        is_in_state = is_in_state_order(order, ship_from_state)
        tax_code = 'Tax' if is_in_state else 'Non'

        # Line items
        for item in order.get('lineItems', []):
            product_name = item.get('productName', 'Product')
            variant_raw = item.get('variantOptions', '')
            variant = parse_variant_options(variant_raw)

            unit_price = item.get('unitPricePaid', {}).get('value', 0)
            unit_price = float(unit_price) if unit_price else 0.0
            quantity = item.get('quantity', 1)
            line_total = quantity * unit_price

            # Map to QB item
            if sku_mapper:
                qb_item = sku_mapper.get_mapping(product_name, variant, unit_price)
            else:
                qb_item = f"{product_name} - {variant}" if variant else product_name

            description = f"{product_name} - {variant}" if variant else product_name

            ws_lines.cell(row=row, column=1, value=order_num)
            ws_lines.cell(row=row, column=2, value=customer)
            ws_lines.cell(row=row, column=3, value=invoice_date)
            ws_lines.cell(row=row, column=4, value=qb_item)
            ws_lines.cell(row=row, column=5, value=description)
            ws_lines.cell(row=row, column=6, value=quantity)
            ws_lines.cell(row=row, column=7, value=unit_price).number_format = currency_fmt
            ws_lines.cell(row=row, column=8, value=line_total).number_format = currency_fmt
            ws_lines.cell(row=row, column=9, value=tax_code)
            row += 1

        # Discount lines
        for disc in order.get('discountLines', []):
            disc_amount = float(disc.get('amount', {}).get('value', 0) or 0)
            if disc_amount <= 0:
                continue

            disc_name = disc.get('name', '') or ''
            disc_promo = disc.get('promoCode', '') or ''

            if 'early access' in disc_name.lower():
                qb_discount_item = '2025 Early Access 10% Off'
                disc_desc = disc_name
            elif disc_promo:
                qb_discount_item = 'Non-inventory Item'
                disc_desc = disc_promo
            else:
                qb_discount_item = 'Non-inventory Item'
                disc_desc = disc_name

            ws_lines.cell(row=row, column=1, value=order_num)
            ws_lines.cell(row=row, column=2, value=customer)
            ws_lines.cell(row=row, column=3, value=invoice_date)
            ws_lines.cell(row=row, column=4, value=qb_discount_item)
            ws_lines.cell(row=row, column=5, value=disc_desc)
            ws_lines.cell(row=row, column=6, value=1)
            ws_lines.cell(row=row, column=7, value=-disc_amount).number_format = currency_fmt
            ws_lines.cell(row=row, column=8, value=-disc_amount).number_format = currency_fmt
            ws_lines.cell(row=row, column=9, value=tax_code)
            row += 1

        # Gift card redemption
        gift_card = order.get('giftCardRedemption', {})
        if gift_card:
            gc_amount = float(gift_card.get('amount', {}).get('value', 0) or 0)
            if gc_amount > 0:
                gc_code = gift_card.get('giftCardCode', 'Gift Card')
                ws_lines.cell(row=row, column=1, value=order_num)
                ws_lines.cell(row=row, column=2, value=customer)
                ws_lines.cell(row=row, column=3, value=invoice_date)
                ws_lines.cell(row=row, column=4, value='Non-inventory Item')
                ws_lines.cell(row=row, column=5, value=f"Gift Card - {gc_code}")
                ws_lines.cell(row=row, column=6, value=1)
                ws_lines.cell(row=row, column=7, value=-gc_amount).number_format = currency_fmt
                ws_lines.cell(row=row, column=8, value=-gc_amount).number_format = currency_fmt
                ws_lines.cell(row=row, column=9, value=tax_code)
                row += 1

        # Freight line
        shipping_total = order.get('shippingTotal', {}).get('value', 0)
        shipping_total = float(shipping_total) if shipping_total else 0.0
        ws_lines.cell(row=row, column=1, value=order_num)
        ws_lines.cell(row=row, column=2, value=customer)
        ws_lines.cell(row=row, column=3, value=invoice_date)
        ws_lines.cell(row=row, column=4, value='Freight')
        ws_lines.cell(row=row, column=5, value='Shipping')
        ws_lines.cell(row=row, column=6, value='')
        ws_lines.cell(row=row, column=7, value=shipping_total).number_format = currency_fmt
        ws_lines.cell(row=row, column=8, value=shipping_total).number_format = currency_fmt
        ws_lines.cell(row=row, column=9, value='')
        row += 1

        # Separator row
        for col_idx in range(1, len(line_headers) + 1):
            ws_lines.cell(row=row, column=col_idx).fill = separator_fill
        row += 1

    ws_lines.freeze_panes = 'A2'
    _auto_fit_columns(ws_lines)

    # =========================================================================
    # Sheet 3: New Customers (if any)
    # =========================================================================
    if new_customers:
        ws_cust = wb.create_sheet("New Customers")

        cust_headers = ['Customer Name', 'Email', 'Phone', 'Bill Address', 'Ship Address', 'Tax Code']
        for col, header in enumerate(cust_headers, 1):
            cell = ws_cust.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for name, info in sorted(new_customers.items()):
            ws_cust.cell(row=row, column=1, value=name)
            ws_cust.cell(row=row, column=2, value=info['email'])
            ws_cust.cell(row=row, column=3, value=info['phone'])
            ws_cust.cell(row=row, column=4, value=info['bill_address'])
            ws_cust.cell(row=row, column=5, value=info['ship_address'])
            ws_cust.cell(row=row, column=6, value=info['tax_code'])
            row += 1

        ws_cust.freeze_panes = 'A2'
        _auto_fit_columns(ws_cust)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _auto_fit_columns(ws, min_width=8, max_width=50):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = min_width
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)
